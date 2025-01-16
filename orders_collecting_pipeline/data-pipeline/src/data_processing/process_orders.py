import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo



def process_order(customer_name, clothing_model, color, size, number):
     # Obtener la fecha actual
    current_date = datetime.now().strftime("%d%m%Y")
    current_time = datetime.now().strftime("%H:%M:%S")
    table_name = 'Tabla' + current_date

    # Validate order data
    if not all([customer_name, clothing_model, color, size, number, current_date, current_time]):
        raise ValueError("Por favor, rellena todos los campos.")

    # Create order data as a dictionary
    order_data = {
        "customer_name": customer_name,
        "clothing_model": clothing_model,
        "color": color,
        "size": size,
        "number": number,
        "hour": current_time
    }

    def write_to_excel(order_data):
        # Define the Excel file path
        excel_file_path = 'D:/orders_collecting_pipeline/data-pipeline/data/orders.xlsx'

        style = style = TableStyleInfo(name="TableStyleMedium5", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        # Convert the order data to a DataFrame
        df = pd.DataFrame([order_data])

        try:
            # Load the existing workbook
             # Load the existing workbook
            book = load_workbook(excel_file_path)
            if current_date not in book.sheetnames:
                book.create_sheet(current_date)
            ws = book[current_date]

            # Append the data to the worksheet
            for row in df.itertuples(index=False, name=None):
                ws.append(row)

            # Update the table range
            if table_name in ws.tables:
                table = ws.tables[table_name]
                table.ref = f"A1:F{ws.max_row}"
            else:
                table = Table(displayName=table_name, ref=f"A1:F{ws.max_row}")
                ws.add_table(table)
                table.tableStyleInfo = style

            # Save the workbook
            book.save(excel_file_path)

        except FileNotFoundError:
            # Create a new workbook and add the data
            with pd.ExcelWriter(excel_file_path, engine='xlsxwriter', mode='w') as writer:

                # Load the workbook and get the ws
                df.to_excel(writer,  startrow=1, index=False, header=False, sheet_name=current_date)
                book = writer.book
                ws = writer.sheets[current_date]               

                # Write the header
                headers = ["nombre", "item", "color", "talla", "cantidad", "hora"]
                column_settings = [{'header': column} for column in headers]
                (max_row, max_col) = df.shape
                
                ws.add_table(0,0,max_row, max_col-1, {"columns": column_settings, "name": table_name, "style": "TableStyleMedium5"})
                
    # Llamar a la función para escribir en el archivo Excel
    write_to_excel(order_data)



